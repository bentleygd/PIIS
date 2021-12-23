from os import remove
from csv import DictWriter

from openpyxl import Workbook

from libs.pii_scan import PIIScanner


class TestFilePIIScan:
    def test_ssn_dashes(self):
        scanner = PIIScanner()
        scan_file = open('test.txt', mode='w+')
        scan_file.write('The data is 123-45-6789')
        scan_file.close()
        test_results = scanner.ssn_scan_file(scan_file.name)
        remove('test.txt')
        assert test_results is True

    def test_ssn_spaces(self):
        scanner = PIIScanner()
        scan_file = open('test.txt', mode='w+')
        scan_file.write('The data is 123 45 6789')
        scan_file.close()
        test_results = scanner.ssn_scan_file(scan_file.name)
        remove('test.txt')
        assert test_results is True

    def test_ssn(self):
        scanner = PIIScanner()
        scan_file = open('test.txt', mode='w+')
        scan_file.write('The data is 123456789')
        scan_file.close()
        test_results = scanner.ssn_scan_file(scan_file.name)
        remove('test.txt')
        assert test_results is True

    def test_ssn_scan_false(self):
        scanner = PIIScanner()
        scan_file = open('test.txt', mode='w+')
        scan_file.write('No numbers')
        scan_file.close()
        test_results = scanner.ssn_scan_file(scan_file.name)
        remove('test.txt')
        assert test_results is False

    def test_ssn_scan_false_phone(self):
        scanner = PIIScanner()
        scan_file = open('test.txt', mode='w+')
        scan_file.write('The phone number is (123)456-7890')
        scan_file.close()
        test_results = scanner.ssn_scan_file(scan_file.name)
        remove('test.txt')
        assert test_results is False


class TestXLSXScan:
    def test_ssn_dashes(self):
        scanner = PIIScanner()
        wb = Workbook()
        sheet_names = ['test', 'temp']
        for name in sheet_names:
            wb.create_sheet(name)
        wb['test']['a3'] = '123-45-6789'
        wb['temp']['a2'] = '123-45-6789'
        wb.save('PyTest.xlsx')
        test_results = scanner.ssn_scan_excel('PyTest.xlsx')
        remove('PyTest.xlsx')
        assert test_results is True

    def test_ssn_spaces(self):
        scanner = PIIScanner()
        wb = Workbook()
        sheet_names = ['test', 'temp']
        for name in sheet_names:
            wb.create_sheet(name)
        wb['test']['a3'] = '123 45 6789'
        wb['temp']['a2'] = '123 45 6789'
        wb.save('PyTest.xlsx')
        test_results = scanner.ssn_scan_excel('PyTest.xlsx')
        remove('PyTest.xlsx')
        assert test_results is True

    def test_ssn(self):
        scanner = PIIScanner()
        wb = Workbook()
        sheet_names = ['test', 'temp']
        for name in sheet_names:
            wb.create_sheet(name)
        wb['test']['a3'] = '123456789'
        wb['temp']['a2'] = '123456789'
        wb.save('PyTest.xlsx')
        test_results = scanner.ssn_scan_excel('PyTest.xlsx')
        remove('PyTest.xlsx')
        assert test_results is True

    def test_ssn_false(self):
        scanner = PIIScanner()
        wb = Workbook()
        sheet_names = ['test', 'temp']
        for name in sheet_names:
            wb.create_sheet(name)
        wb['test']['a3'] = '1234567890'
        wb['temp']['a2'] = 'Not a number'
        wb.save('PyTest.xlsx')
        test_results = scanner.ssn_scan_excel('PyTest.xlsx')
        remove('PyTest.xlsx')
        assert test_results is False


class TestCSVPIIScan:
    def test_ssn_dashes(self):
        scanner = PIIScanner()
        scan_file = open('test.csv', mode='w+', newline='')
        fieldnames = ['Name', 'SSN']
        csv_writer = DictWriter(scan_file, fieldnames=fieldnames)
        csv_writer.writeheader()
        csv_writer.writerow({
            'Name': 'John Smith',
            'SSN': '123-45-6789'
            })
        scan_file.close()
        test_results = scanner.ssn_scan_csv(scan_file.name)
        remove('test.csv')
        assert test_results is True

    def test_ssn_spaces(self):
        scanner = PIIScanner()
        scan_file = open('test.csv', mode='w+', newline='')
        fieldnames = ['Name', 'SSN']
        csv_writer = DictWriter(scan_file, fieldnames=fieldnames)
        csv_writer.writeheader()
        csv_writer.writerow({
            'Name': 'John Smith',
            'SSN': '123 45 6789'
            })
        scan_file.close()
        test_results = scanner.ssn_scan_csv(scan_file.name)
        remove('test.csv')
        assert test_results is True

    def test_ssn(self):
        scanner = PIIScanner()
        scan_file = open('test.csv', mode='w+', newline='')
        fieldnames = ['Name', 'SSN']
        csv_writer = DictWriter(scan_file, fieldnames=fieldnames)
        csv_writer.writeheader()
        csv_writer.writerow({
            'Name': 'John Smith',
            'SSN': '123456789'
            })
        scan_file.close()
        test_results = scanner.ssn_scan_csv(scan_file.name)
        remove('test.csv')
        assert test_results is True

    def test_ssn_scan_false(self):
        scanner = PIIScanner()
        scan_file = open('test.csv', mode='w+', newline='')
        fieldnames = ['Name', 'Not SSN']
        csv_writer = DictWriter(scan_file, fieldnames=fieldnames)
        csv_writer.writeheader()
        csv_writer.writerow({
            'Name': 'John Smith',
            'Not SSN': '12345678912309182310293'
            })
        scan_file.close()
        test_results = scanner.ssn_scan_csv(scan_file.name)
        remove('test.csv')
        assert test_results is False

    def test_ssn_scan_false_phone(self):
        scanner = PIIScanner()
        scan_file = open('test.csv', mode='w+', newline='')
        fieldnames = ['Name', 'Phone']
        csv_writer = DictWriter(scan_file, fieldnames=fieldnames)
        csv_writer.writeheader()
        csv_writer.writerow({
            'Name': 'John Smith',
            'Phone': '(888)555-1212'
            })
        scan_file.close()
        test_results = scanner.ssn_scan_csv(scan_file.name)
        remove('test.csv')
        assert test_results is False
